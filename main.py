import zipfile
import os
import re
import sys
from PIL import Image
from fpdf import FPDF
import pytesseract
import fitz
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from typing import List, Dict

pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

def convert_zip_to_pdf(zip_path: str, output_pdf: str) -> str:
    import shutil
    import time
    extract_path = "temp_extracted_images"
    if os.path.exists(extract_path):
        try:
            shutil.rmtree(extract_path)
        except:
            pass
        time.sleep(0.5)
    with zipfile.ZipFile(zip_path, 'r') as zip_ref:
        zip_ref.extractall(extract_path)
    image_files = sorted([
        os.path.join(extract_path, f)
        for f in os.listdir(extract_path)
        if f.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp', '.tiff', '.gif'))
    ])
    if not image_files:
        raise ValueError("No image files found in ZIP")
    pdf = FPDF()
    for image in image_files:
        try:
            with Image.open(image) as img:
                width, height = img.size
                width_mm, height_mm = width * 0.264583, height * 0.264583
                pdf.add_page()
                pdf.image(image, 0, 0, width_mm, height_mm)
        except Exception as e:
            print(f"Warning: Could not process image {image}: {e}")
    pdf.output(output_pdf)
    for _ in range(3):
        try:
            shutil.rmtree(extract_path)
            break
        except:
            time.sleep(1)
    return output_pdf

def extract_text_from_pdf(pdf_path: str, output_text: str) -> str:
    doc = fitz.open(pdf_path)
    text_data = ""
    for page_num, page in enumerate(doc):
        text = page.get_text()
        if not text.strip():
            try:
                pix = page.get_pixmap()
                from io import BytesIO
                img = Image.open(BytesIO(pix.tobytes("png")))
                text = pytesseract.image_to_string(img)
            except Exception as e:
                text = f"[OCR Error: {str(e)}]"
                print(f"OCR failed for page {page_num + 1}: {e}")
        text_data += f"\n--- Page {page_num + 1} ---\n{text}"
    doc.close()
    with open(output_text, "w", encoding="utf-8") as f:
        f.write(text_data)
    return output_text

def extract_contacts_from_text(text: str) -> List[Dict]:
    contacts = []
    page_pattern = r'--- Page \d+ ---'
    pages = re.split(page_pattern, text)
    for page_content in pages:
        if not page_content.strip():
            continue
        contact = extract_contact_from_page(page_content.strip())
        if contact and contact.get('Name', '').strip():
            contacts.append(contact)
    return contacts

def extract_contact_from_page(page_content: str) -> Dict:
    lines = [line.strip() for line in page_content.split('\n') if line.strip()]
    contact = {
        'Name': 'N/A',
        'Title': 'N/A',
        'Company': 'N/A',
        'Email': 'N/A',
        'Mobile Phone': 'N/A',
        'Direct Phone': 'N/A',
        'HQ Phone': 'N/A',
        'Location': 'N/A'
    }
    if not lines:
        return contact
    for line in lines:
        cleaned_line = clean_text(line)
        if is_person_name(cleaned_line):
            contact['Name'] = cleaned_line
            break
    title_keywords = [
        'ceo', 'cto', 'cfo', 'coo', 'founder', 'co-founder', 'director', 'managing director',
        'executive director', 'senior director', 'vice president', 'vp', 'president',
        'manager', 'senior manager', 'head', 'lead', 'chairman', 'chairwoman'
    ]
    for line in lines:
        line_lower = line.lower()
        if any(keyword in line_lower for keyword in title_keywords):
            contact['Title'] = clean_text(line)
            break
    company_indicators = [
        'properties', 'real estate', 'group', 'company', 'corp', 'inc',
        'llc', 'ltd', 'development', 'investments', 'solutions',
        'technologies', 'services', 'international', 'global', 'brokers',
        'realty', 'developers', 'management', 'holdings', 'ventures',
        'hub', 'center', 'centre', 'mall', 'tower', 'square', 'homes'
    ]
    for line in lines:
        line_clean = clean_text(line).lower()
        if any(indicator in line_clean for indicator in company_indicators):
            cleaned = clean_text(line)
            if len(cleaned) > 3 and not is_person_name(cleaned):
                contact['Company'] = cleaned
                break
    email_pattern = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
    for line in lines:
        email_match = re.search(email_pattern, line)
        if email_match:
            contact['Email'] = email_match.group()
            break
    for line in lines:
        mobile_match = re.search(r'\+\d{3}\s*\d{2}\s*\d{3}\s*\d{4}\s*\(M\)', line)
        if mobile_match:
            contact['Mobile Phone'] = clean_phone_number(mobile_match.group())
        direct_match = re.search(r'\+\d{3}\s*\d{1,2}\s*\d{3}\s*\d{4}\s*\(D\)', line)
        if direct_match:
            contact['Direct Phone'] = clean_phone_number(direct_match.group())
        hq_match = re.search(r'\+\d{3}\s*\d{8,9}\s*\(HQ\)', line)
        if hq_match:
            contact['HQ Phone'] = clean_phone_number(hq_match.group())
    location_lines = []
    in_location_section = False
    for line in lines:
        if line.lower().startswith('location') or line.lower().startswith('local'):
            in_location_section = True
            continue
        elif in_location_section:
            if any(section in line.lower() for section in ['crm', 'main contact', 'additional contact']):
                break
            if line and not line.startswith('---'):
                location_lines.append(line)
    if location_lines:
        contact['Location'] = clean_location(' '.join(location_lines))
    return contact

def is_person_name(text: str) -> bool:
    if not text:
        return False
    words = text.split()
    if len(words) < 2:
        return False
    title_keywords = ['ceo', 'cto', 'cfo', 'director', 'manager', 'president', 'founder', 'vice', 'senior']
    text_lower = text.lower()
    if any(keyword in text_lower for keyword in title_keywords):
        return False
    company_indicators = ['properties', 'real estate', 'group', 'company', 'corp', 'inc', 'llc', 'ltd']
    if any(indicator in text_lower for indicator in company_indicators):
        return False
    return all(word[0].isupper() for word in words if word and word.isalpha())

def clean_text(text: str) -> str:
    if not text:
        return ""
    text = re.sub(r'[@#$%^&*()_+=\[\]{}|\\:";\'<>?,./`~]', ' ', text)
    text = re.sub(r'\b[a-z]\s+(?=[A-Z])', ' ', text)
    text = re.sub(r'\s+', ' ', text)
    text = re.sub(r'^\d+\s*', '', text)
    return text.strip()

def clean_phone_number(phone: str) -> str:
    numbers = re.findall(r'\d+', phone)
    if numbers:
        return '+' + ''.join(numbers)
    return phone

def clean_location(location: str) -> str:
    location = re.sub(r'[^\w\s,.-]', '', location)
    location = re.sub(r'\s+', ' ', location).strip()
    location_fixes = {
        'dubai dubai': 'Dubai',
        'united arab emirates': 'United Arab Emirates',
        'uae': 'UAE'
    }
    location_lower = location.lower()
    for old, new in location_fixes.items():
        location_lower = location_lower.replace(old, new)
    return location_lower.title()

def create_excel_file(contacts: List[Dict], output_file: str) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Contacts"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    headers = ['Name', 'Title', 'Company', 'Email', 'Mobile Phone', 'Direct Phone', 'HQ Phone', 'Location']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    for row, contact in enumerate(contacts, 2):
        for col, header in enumerate(headers, 1):
            value = contact.get(header, 'N/A')
            if not value or value.strip() == '':
                value = 'N/A'
            ws.cell(row=row, column=col, value=value)
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    wb.save(output_file)
    return output_file

def process_file(input_file: str, output_excel: str = None) -> str:
    if not os.path.exists(input_file):
        raise FileNotFoundError(f"Input file not found: {input_file}")
    base_name = os.path.splitext(os.path.basename(input_file))[0]
    if not output_excel:
        output_excel = f"{base_name}_contacts.xlsx"
    text_file = f"{base_name}_text.txt"
    file_ext = os.path.splitext(input_file)[1].lower()
    if file_ext == '.zip':
        print(f"📦 Processing ZIP file: {input_file}")
        temp_pdf = f"temp_{base_name}.pdf"
        print("Step 1: Converting ZIP to PDF...")
        convert_zip_to_pdf(input_file, temp_pdf)
        print("Step 2: Extracting text from PDF...")
        extract_text_from_pdf(temp_pdf, text_file)
        print(f"✓ Text saved: {text_file}")
        os.remove(temp_pdf)
    elif file_ext == '.pdf':
        print(f"📄 Processing PDF file: {input_file}")
        print("Step 1: Extracting text from PDF...")
        extract_text_from_pdf(input_file, text_file)
        print(f"✓ Text saved: {text_file}")
    else:
        raise ValueError(f"Unsupported file type: {file_ext}. Only ZIP and PDF files are supported.")
    print("Step 2: Extracting contact information...")
    with open(text_file, 'r', encoding='utf-8') as f:
        text_content = f.read()
    contacts = extract_contacts_from_text(text_content)
    print(f"✓ Found {len(contacts)} contacts")
    print("Step 3: Creating Excel file...")
    excel_path = create_excel_file(contacts, output_excel)
    print(f"✓ Excel file created: {excel_path}")
    print(f"\n📊 Summary:")
    print(f"   • Input file: {input_file}")
    print(f"   • Text file: {text_file}")
    print(f"   • Excel file: {excel_path}")
    print(f"   • Contacts extracted: {len(contacts)}")
    return excel_path

def main():
    if len(sys.argv) < 2:
        print("Usage: python dynamic_contact_processor.py <input_file> [output_excel]")
        print("       input_file: ZIP or PDF file to process")
        print("       output_excel: Optional output Excel filename")
        print("\nExample:")
        print("       python dynamic_contact_processor.py contacts.zip")
        print("       python dynamic_contact_processor.py document.pdf my_contacts.xlsx")
        return
    input_file = sys.argv[1]
    output_excel = sys.argv[2] if len(sys.argv) > 2 else None
    try:
        result = process_file(input_file, output_excel)
        print(f"\n✅ Processing complete! Output saved to: {result}")
    except Exception as e:
        print(f"\n❌ Error: {e}")

if __name__ == "__main__":
    main()
